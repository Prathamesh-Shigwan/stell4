# accounts/views.py
import openpyxl
from django.db.models import Sum, Count, Q
from django.views.decorators.http import require_POST
from products.models import MainCategory, SubCategory, Product, CartItem, Order, OrderItem, \
    ExtraImages, ProductVariant, VariantExtraImage, Cart, CartItem, \
    Wishlist, ShippingAddress, BillingAddress, BannerImage, About, SiteSettings, DiscountCode, VariantSizeOption, \
    MediaLibrary, SiteContent, Brand
from accounts.models import User
from accounts.forms import UserRoleUpdateForm, ProfileForm, UserUpdateForm
from django.http import FileResponse
from django.contrib.auth import authenticate, login
from django.contrib.auth.forms import AuthenticationForm
from blog.models import Blog
from django.contrib.admin.views.decorators import staff_member_required
from datetime import datetime, timedelta, date, timezone, time
from django.http import JsonResponse
from django.utils.timezone import make_aware, localtime
from django.contrib.admin.models import LogEntry
from django.db.models.functions import TruncDay, TruncMonth
from django.db.models.functions import ExtractWeek, ExtractYear
from decimal import Decimal
import openpyxl
from django.http import HttpResponse, JsonResponse
from openpyxl.styles import Font, Alignment, PatternFill
from decimal import Decimal # Make sure this is also there now
from django.db.models import Prefetch, Sum, Count # Add Prefetch here!
from django.utils.timezone import now, make_aware, localtime
import json
from django.utils.timezone import now
from django.db import models  # Import the models module
from custom_admin.utils import get_recent_actions_ut  # Ensure you have this function to fetch recent actions
from custom_admin.forms import DateRangeForm, MainCategoryForm, SubCategoryForm, \
    ProductForm, ProductVariantForm, VariantExtraImageForm, \
    OrderForm, CartForm, WishlistForm, VariantSizeOptionFormSet, VariantExtraImageFormSet, \
    BlogForm, BannerImageForm, CustomerForm, ProfileForm, AboutForm, SiteSettingsForm, DiscountCodeForm, \
    SiteContentForm, BrandForm, CustomAdminAuthenticationForm
from django.contrib.auth.decorators import user_passes_test
from django.shortcuts import get_object_or_404, redirect, render
from django.forms import modelformset_factory
from django.db import transaction
from django.forms import inlineformset_factory
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import HttpResponse
from PIL import Image, ImageDraw, ImageFont
import barcode
from barcode.writer import ImageWriter
import io
import pandas as pd
import os
from django.conf import settings
from django.contrib import messages
from django.core.files import File
from django.core.files.images import ImageFile
from django.utils.text import slugify
import json
import openpyxl
from django.http import HttpResponse
from reportlab.lib.units import mm
from reportlab.pdfgen import canvas
from io import BytesIO
from products.models import ProductVariant, VariantSizeOption
from reportlab.graphics.barcode import code128
import zipfile
from collections import defaultdict
from django.utils.timezone import localdate
from django.contrib import messages
from django.core.mail import send_mail
from django.template.loader import render_to_string
from django.contrib.auth.decorators import login_required


def is_admin(user):
    return user.is_superuser


@login_required(login_url='custom_admin:admin_login')
def dashboard(request):
    # The decorator above handles the login check, so the manual 'if not request.user.is_authenticated' is no longer needed.

    # Check if the logged-in user has a role permitted to see the dashboard.
    # We can expand this to include other relevant staff roles.
    allowed_roles = ['admin', 'finance_manager', 'inventory_manager', 'content_manager', 'staff']
    if request.user.role in allowed_roles:

        # --- All the logic to gather data for the dashboard remains the same ---
        today = now()
        default_start_date = today.replace(day=1).date()
        default_end_date = today.date()

        form = DateRangeForm(request.GET or None)
        start_date_obj = default_start_date
        end_date_obj = default_end_date

        if form.is_valid():
            start_date_obj = form.cleaned_data.get('start_date') or default_start_date
            end_date_obj = form.cleaned_data.get('end_date') or default_end_date

        start_date_aware = make_aware(datetime.combine(start_date_obj, time.min))
        end_date_aware = make_aware(datetime.combine(end_date_obj, time.max))

        orders_in_range = Order.objects.filter(created_at__range=(start_date_aware, end_date_aware))

        total_sales = orders_in_range.filter(status='completed').aggregate(Sum('total'))['total__sum'] or 0.00
        total_orders = orders_in_range.count()

        payment_methods_data = orders_in_range.values('payment_method').annotate(count=Count('id'))

        prepaid_orders = 0
        cod_orders = 0
        for method in payment_methods_data:
            payment_method_str = (method.get('payment_method') or '').lower()
            if 'cash' in payment_method_str:
                cod_orders += method.get('count', 0)
            else:
                prepaid_orders += method.get('count', 0)

        total_products = Product.objects.count()
        total_variants = ProductVariant.objects.count()
        products_in_cart = CartItem.objects.aggregate(total=Sum('quantity'))['total'] or 0
        canceled_orders = orders_in_range.filter(status='cancelled').count()
        returned_orders = orders_in_range.filter(status='returned').count()
        replaced_orders = orders_in_range.filter(status='replaced').count()
        total_customers = User.objects.filter(is_staff=False).count()
        total_blogs = Blog.objects.count()
        recent_orders = Order.objects.select_related('user').order_by('-created_at')[:10]
        recent_actions = get_recent_actions(request)

        context = {
            'form': form,
            'total_sales': total_sales,
            'total_orders': total_orders,
            'prepaid_orders': prepaid_orders,
            'cod_orders': cod_orders,
            'payment_methods': payment_methods_data,
            'total_products': total_products,
            'products_in_cart': products_in_cart,
            'canceled_orders': canceled_orders,
            'returned_orders': returned_orders,
            'replaced_orders': replaced_orders,
            'total_customers': total_customers,
            'total_blogs': total_blogs,
            'recent_orders': recent_orders,
            'recent_actions': recent_actions,
            'total_variants': total_variants,
        }
        return render(request, 'custom_admin/dashboard.html', context)

    else:
        # If the user is logged in but doesn't have a permitted role, show an access denied message.
        return render(request, 'custom_admin/welcome.html', {
            'message': 'Access Denied: You do not have the required permissions to view the dashboard.'
        })

@staff_member_required
def export_orders_excel(request):
    # Define the headers for directly inheritable/calculable fields
    headers = [
        "Payment_Method", "Currency", "Total_Amount",
        "Postpaid_Amount", "Prepaid_Amount", "Line_Item_Price_At_Order", "Order_Discount_Amount",
        "Shipping_Case", "Order_Code", "Applicable_Tax_Rate",
        "Igst_Amount", "Cgst_Amount", "Sgst_Amount",
        "Gender", "Product_Name", "Order_Created_Date",
        "Customer_Name", "Customer_PinCode", "Customer_State", "Customer_Address",
        "Customer_Email", "Customer_Contact_No", # <--- ADDED HEADERS HERE
        "Igst_Rate", "Cgst_Rate", "Sgst_Rate",
        "Line_Item_Taxable_Amount",
        "Shipping_Amount", "Coupon_Code_Used",
        "Seller_State_Code", "SKU_Code"
    ]

    # Prepare data for the Excel report in a list of lists format
    report_data = []

    # Fetch SiteSettings once
    site_settings = SiteSettings.objects.first()
    shipping_charge = site_settings.shipping_charge if site_settings else Decimal('0.00')
    TAX_RATE = Decimal('0.18')  # As seen in your invoice logic

    # Seller's state is constant (Mumbai, Maharashtra, India)
    SELLER_STATE_CODE = "MH"

    # Fetch orders with related data to minimize queries
    orders = Order.objects.select_related('user').prefetch_related(
        Prefetch('items', queryset=OrderItem.objects.select_related(
            'product',
            'product_variant',
        ))
    ).order_by('-created_at')

    for order in orders:
        is_maharashtra = order.shipping_state.strip().lower() == "maharashtra".lower()
        shipping_case = "Intrastate" if is_maharashtra else "Interstate"

        grand_total_to_pay = order.total + shipping_charge
        postpaid_amount = grand_total_to_pay if order.payment_method == "Cash on Delivery" else Decimal('0.00')
        prepaid_amount = grand_total_to_pay if order.payment_method != "Cash on Delivery" else Decimal('0.00')
        order_discount_amount = order.discount if order.discount else Decimal('0.00')

        for item in order.items.all():
            product = item.product
            product_variant = item.product_variant

            product_name_base = product.name if product else 'N/A'
            sku_code = product.sku if product else 'N/A'

            gender = 'N/A'
            if product_variant and product_variant.gender:
                gender = product_variant.gender
            elif product and product.variants.first() and product.variants.first().gender:
                gender = product.variants.first().gender

            variant_color = product_variant.color if product_variant else 'N/A'
            variant_size = item.selected_size if item.selected_size else 'N/A'

            parts = [product_name_base]
            if variant_color and variant_color != 'N/A':
                parts.append(variant_color)
            if variant_size and variant_size != 'N/A':
                parts.append(variant_size)
            detailed_product_name = "_".join(parts).replace(' ', '_').lower()

            line_item_total_price = item.quantity * item.price
            line_item_taxable_amount = line_item_total_price / (Decimal('1') + TAX_RATE)
            item_tax_amount_total = line_item_total_price - line_item_taxable_amount

            igst_amount = Decimal('0.00')
            cgst_amount = Decimal('0.00')
            sgst_amount = Decimal('0.00')

            igst_rate = Decimal('0.00')
            cgst_rate = Decimal('0.00')
            sgst_rate = Decimal('0.00')

            if shipping_case == "Interstate":
                igst_amount = item_tax_amount_total
                igst_rate = TAX_RATE
            else:
                cgst_amount = item_tax_amount_total / 2
                sgst_amount = item_tax_amount_total / 2
                cgst_rate = TAX_RATE / 2
                sgst_rate = TAX_RATE / 2

            customer_address = f"{order.shipping_address1}, {order.shipping_address2 or ''}"
            if order.shipping_city: customer_address += f", {order.shipping_city}"
            if order.shipping_state: customer_address += f", {order.shipping_state}"
            if order.shipping_zipcode: customer_address += f", {order.shipping_zipcode}"
            if order.shipping_country: customer_address += f", {order.shipping_country}"

            report_data.append([
                order.payment_method,
                "INR",
                float(grand_total_to_pay),
                float(postpaid_amount),
                float(prepaid_amount),
                float(line_item_total_price),
                float(order_discount_amount),
                shipping_case,
                order.order_id,
                float(TAX_RATE),
                float(igst_amount),
                float(cgst_amount),
                float(sgst_amount),
                gender,
                detailed_product_name,
                order.created_at.strftime('%Y-%m-%d %H:%M:%S'),
                order.shipping_full_name,
                order.shipping_zipcode,
                order.shipping_state,
                customer_address,
                order.shipping_email,      # <--- ADDED CUSTOMER EMAIL
                order.shipping_phone,      # <--- ADDED CUSTOMER PHONE
                float(igst_rate),
                float(cgst_rate),
                float(sgst_rate),
                float(line_item_taxable_amount),
                float(shipping_charge),
                order.discount_code if order.discount_code else "",
                SELLER_STATE_CODE,
                sku_code
            ])

    return generate_excel_report("Orders Report(Product)", headers=headers, report_data=report_data)



@staff_member_required
def export_orders_summary_excel(request):
    """
    Generates an Excel report with one row per order,
    providing aggregated order-level sales data.
    """
    headers = [
        "Payment_Method",
        "Currency",
        "Order_Total_Including_Shipping",
        "Postpaid_Amount",
        "Prepaid_Amount",
        "Order_Subtotal_Before_Tax_Shipping_Discount",
        "Order_Discount_Amount",
        "Shipping_Case",
        "Order_Code",
        "Applicable_Tax_Rate_Per_Item",
        "Total_IGST_Amount_Order",
        "Total_CGST_Amount_Order",
        "Total_SGST_Amount_Order",
        "Order_Created_Date",
        "Customer_Name",
        "Customer_PinCode",
        "Customer_State",
        "Customer_Address",
        "Customer_Email", "Customer_Contact_No", # <--- ADDED HEADERS HERE
        "IGST_Rate_Per_Item",
        "CGST_Rate_Per_Item",
        "SGST_Rate_Per_Item",
        "Order_Taxable_Amount",
        "Shipping_Amount",
        "Coupon_Code_Used",
        "Seller_State_Code",
        "Order_Status",
    ]

    report_data = []

    site_settings = SiteSettings.objects.first()
    shipping_charge = Decimal(str(site_settings.shipping_charge)) if site_settings else Decimal('0.00')
    TAX_RATE = Decimal('0.18')

    SELLER_STATE_CODE = "MH"

    orders = Order.objects.select_related('user').prefetch_related('items').order_by('-created_at')

    for order in orders:
        is_maharashtra = order.shipping_state.strip().lower() == "maharashtra".lower()
        shipping_case = "Intrastate" if is_maharashtra else "Interstate"

        grand_total_to_pay = order.total + shipping_charge
        postpaid_amount = grand_total_to_pay if order.payment_method == "Cash on Delivery" else Decimal('0.00')
        prepaid_amount = grand_total_to_pay if order.payment_method != "Cash on Delivery" else Decimal('0.00')
        order_discount_amount = order.discount if order.discount else Decimal('0.00')

        order_subtotal_before_tax_shipping_discount = Decimal('0.00')
        order_total_taxable_amount = Decimal('0.00')

        for item in order.items.all():
            line_item_total_price = item.quantity * item.price
            item_taxable_amount = line_item_total_price / (Decimal('1') + TAX_RATE)

            order_subtotal_before_tax_shipping_discount += line_item_total_price
            order_total_taxable_amount += item_taxable_amount

        order_total_tax_amount = order_subtotal_before_tax_shipping_discount - order_total_taxable_amount

        igst_amount_order = Decimal('0.00')
        cgst_amount_order = Decimal('0.00')
        sgst_amount_order = Decimal('0.00')

        igst_rate_per_item = Decimal('0.00')
        cgst_rate_per_item = Decimal('0.00')
        sgst_rate_per_item = Decimal('0.00')

        if shipping_case == "Interstate":
            igst_amount_order = order_total_tax_amount
            igst_rate_per_item = TAX_RATE
        else:
            cgst_amount_order = order_total_tax_amount / 2
            sgst_amount_order = order_total_tax_amount / 2
            cgst_rate_per_item = TAX_RATE / 2
            sgst_rate_per_item = TAX_RATE / 2

        customer_address_parts = [order.shipping_address1]
        if order.shipping_address2:
            customer_address_parts.append(order.shipping_address2)
        if order.shipping_city:
            customer_address_parts.append(order.shipping_city)
        if order.shipping_state:
            customer_address_parts.append(order.shipping_state)
        if order.shipping_zipcode:
            customer_address_parts.append(order.shipping_zipcode)
        if order.shipping_country:
            customer_address_parts.append(order.shipping_country)
        customer_address = ", ".join(filter(None, customer_address_parts))


        report_data.append([
            order.payment_method,
            "INR",
            float(grand_total_to_pay),
            float(postpaid_amount),
            float(prepaid_amount),
            float(order_subtotal_before_tax_shipping_discount),
            float(order_discount_amount),
            shipping_case,
            order.order_id,
            float(TAX_RATE),
            float(igst_amount_order),
            float(cgst_amount_order),
            float(sgst_amount_order),
            order.created_at.strftime('%Y-%m-%d %H:%M:%S'),
            order.shipping_full_name,
            order.shipping_zipcode,
            order.shipping_state,
            customer_address,
            order.shipping_email,      # <--- ADDED CUSTOMER EMAIL
            order.shipping_phone,      # <--- ADDED CUSTOMER PHONE
            float(igst_rate_per_item),
            float(cgst_rate_per_item),
            float(sgst_rate_per_item),
            float(order_total_taxable_amount),
            float(shipping_charge),
            order.discount_code if order.discount_code else "",
            SELLER_STATE_CODE,
            order.status,
        ])
    return generate_excel_report("Order Report(Order)", headers=headers, report_data=report_data)

def generate_excel_report(report_title, labels=None, data=None, headers=None, report_data=None):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = report_title

    # Add the main report title at the top
    ws.append([report_title])
    ws.append([]) # Add a blank row for spacing

    if headers and report_data is not None: # Case for multi-column detailed report
        ws.append(headers) # Add the column headers
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
        for cell in ws[ws.max_row]: # Apply style to the just added header row
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        for row in report_data:
            ws.append(row)

    elif labels is not None and data is not None: # Case for 2-column statistical report
        ws.append(['Date/Period', 'Value']) # Add default headers for statistical report
        for label, value in zip(labels, data):
            ws.append([label, value])

    else:
        # Handle cases where neither format is correctly provided
        ws.append(["Error: No data format provided"])

    # Optional: Adjust column widths for better readability (applies to all scenarios)
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if cell.value is not None:
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width


    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"{report_title.replace(' ', '_').lower()}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response

def weekly_order_data(request):
    current_year = now().year

    orders = (
        Order.objects
        .annotate(
            week=ExtractWeek('created_at'),
            year=ExtractYear('created_at')
        )
        .filter(year=current_year)
        .values('week')
        .annotate(total=Count('id'))
        .order_by('week')
    )

    labels = []
    data = []

    for entry in orders:
        labels.append(f"Week {entry['week']}")
        data.append(entry['total'])

    return JsonResponse({'labels': labels, 'data': data})


def daily_sales_data(request):
    today = now().date()
    last_7_days = [today - timedelta(days=i) for i in range(6, -1, -1)]
    labels = [day.strftime("%A") for day in last_7_days]
    data = []

    for day in last_7_days:
        start = make_aware(datetime.datetime.combine(day, datetime.datetime.min.time()))
        end = make_aware(datetime.datetime.combine(day, datetime.datetime.max.time()))
        total = Order.objects.filter(status='completed', created_at__range=(start, end)).aggregate(Sum('total'))['total__sum'] or 0
        data.append(total)

    return JsonResponse({'labels': labels, 'data': data})



def monthly_sales_data(request):
    current_year = now().year
    labels = ["January", "February", "March", "April", "May", "June",
              "July", "August", "September", "October", "November", "December"]
    data = {month: 0 for month in labels}

    orders = Order.objects.filter(created_at__year=current_year) \
                .annotate(month=TruncMonth('created_at')) \
                .values('month') \
                .annotate(total=Sum('total'))

    for order in orders:
        month_label = order['month'].strftime("%B")
        data[month_label] = float(order['total'])

    return JsonResponse({'labels': labels, 'data': list(data.values())})


def download_weekly_order_report(request):
    response = weekly_order_data(request)
    data_response = json.loads(response.content)
    # Corrected: Pass the report title as the first argument,
    # then labels and data for the statistical report
    return generate_excel_report('Weekly Order Report', labels=data_response['labels'], data=data_response['data'])

def download_daily_sales_report(request):
    response = daily_sales_data(request)
    data_response = json.loads(response.content)
    return generate_excel_report('Daily Sales Report', labels=data_response['labels'], data=data_response['data'])

def download_monthly_sales_report(request): # <--- This is the function you are calling
    response = monthly_sales_data(request) # <--- This calls the data generation function
    data_response = json.loads(response.content)
    # Corrected based on previous discussion
    return generate_excel_report('Monthly Sales Report', labels=data_response['labels'], data=list(data_response['data']))

def get_recent_actions(request):
    # Fetch the 10 most recent log entries
    recent_actions = LogEntry.objects.select_related('content_type', 'user').order_by('-action_time')[:10]

    # Prepare the data for rendering
    actions = []
    for entry in recent_actions:
        actions.append({
            'object_repr': entry.object_repr,  # Object name
            'action_flag': entry.get_action_flag_display(),  # Action type (Add, Change, Delete)
            'content_type': entry.content_type.name,  # Content type (e.g., Order, Product)
            'user': entry.user.username,  # User who performed the action
            'action_time': localtime(entry.action_time),  # Localized action time
        })

    return actions


@staff_member_required
def main_categories_view(request):
    main_categories = MainCategory.objects.all()

    # --- Search Filter ---
    query = request.GET.get('q')
    if query:
        main_categories = main_categories.filter(
            Q(title__icontains=query) | Q(price__icontains=query)
        )

    # --- Sort Filter ---
    sort_by = request.GET.get('sort_by', 'title') # Default sort by title
    order = request.GET.get('order', 'asc') # Default order ascending

    if sort_by in ['title', 'price']: # Ensure we only sort by valid fields
        if order == 'desc':
            main_categories = main_categories.order_by(f'-{sort_by}')
        else:
            main_categories = main_categories.order_by(sort_by)

    context = {
        'main_categories': main_categories,
        'query': query, # Pass the query back to the template for display
        'sort_by': sort_by, # Pass current sort_by to template
        'order': order,   # Pass current order to template
    }
    return render(request, 'custom_admin/products/main_categories.html', context)


def add_main_category_view(request):
    if request.method == 'POST':
        form = MainCategoryForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            return redirect('custom_admin:main_categories')
    else:
        form = MainCategoryForm()
    return render(request, 'custom_admin/products/add_main_category.html', {'form': form})


def edit_main_category_view(request, pk):
    category = get_object_or_404(MainCategory, pk=pk)
    if request.method == 'POST':
        form = MainCategoryForm(request.POST, request.FILES, instance=category)
        if form.is_valid():
            form.save()
            return redirect('custom_admin:main_categories')
    else:
        form = MainCategoryForm(instance=category)
    return render(request, 'custom_admin/products/edit_main_category.html', {'form': form, 'category': category})


def delete_main_category_view(request, pk):
    category = get_object_or_404(MainCategory, pk=pk)
    if request.method == 'POST':
        category.delete()
        return redirect('custom_admin:main_categories')
    return render(request, 'custom_admin/products/delete_main_category.html', {'category': category})


@staff_member_required
def subcategories_view(request):
    subcategories = SubCategory.objects.all()

    # --- Search Filter ---
    query = request.GET.get('q')
    if query:
        subcategories = subcategories.filter(
            Q(title__icontains=query) |
            Q(main_category__title__icontains=query) | # Search by related MainCategory title
            Q(price__icontains=query)
        )

    # --- Main Category Filter (Dropdown) ---
    main_category_filter = request.GET.get('main_category')
    if main_category_filter:
        subcategories = subcategories.filter(main_category__id=main_category_filter)

    # --- Sort Filter ---
    sort_by = request.GET.get('sort_by', 'title') # Default sort by title
    order = request.GET.get('order', 'asc') # Default order ascending

    if sort_by in ['title', 'main_category__title', 'price']: # Ensure we only sort by valid fields
        if order == 'desc':
            subcategories = subcategories.order_by(f'-{sort_by}')
        else:
            subcategories = subcategories.order_by(sort_by)

    # Get all main categories for the dropdown filter
    main_categories_for_filter = MainCategory.objects.all().order_by('title')

    context = {
        'subcategories': subcategories,
        'query': query, # Pass the query back to the template for display
        'sort_by': sort_by, # Pass current sort_by to template
        'order': order,   # Pass current order to template
        'main_category_filter': main_category_filter, # Pass current main_category filter to template
        'main_categories_for_filter': main_categories_for_filter, # Pass all main categories for dropdown
    }
    return render(request, 'custom_admin/products/subcategories.html', context)


def add_subcategory_view(request):
    if request.method == 'POST':
        form = SubCategoryForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            return redirect('custom_admin:subcategories')
    else:
        form = SubCategoryForm()
    return render(request, 'custom_admin/products/add_subcategory.html', {'form': form})


def edit_subcategory_view(request, pk):
    subcategory = get_object_or_404(SubCategory, pk=pk)
    if request.method == 'POST':
        form = SubCategoryForm(request.POST, request.FILES, instance=subcategory)
        if form.is_valid():
            form.save()
            return redirect('custom_admin:subcategories')
    else:
        form = SubCategoryForm(instance=subcategory)
    return render(request, 'custom_admin/products/edit_subcategory.html', {'form': form, 'subcategory': subcategory})


def delete_subcategory_view(request, pk):
    subcategory = get_object_or_404(SubCategory, pk=pk)
    if request.method == 'POST':
        subcategory.delete()
        return redirect('custom_admin:subcategories')
    return render(request, 'custom_admin/products/delete_subcategory.html', {'subcategory': subcategory})



@staff_member_required
def upload_main_categories_view(request):
    if request.method == 'POST' and request.FILES.get('excel_file'):
        excel_file = request.FILES['excel_file']
        try:
            df = pd.read_excel(excel_file)
        except Exception as e:
            messages.error(request, f"Error reading Excel file: {e}")
            return redirect('custom_admin:main_categories')

        created_count = 0
        for _, row in df.iterrows():
            title = str(row.get('Title')).strip()
            price = row.get('Price') or 0.0
            image_name = str(row.get('Image')).strip()

            # Look for image in MediaLibrary
            media_file = MediaLibrary.objects.filter(name__iexact=image_name).first()
            if media_file:
                image = ImageFile(media_file.file, name=media_file.file.name)
            else:
                # Fallback to default image
                default_path = os.path.join(settings.MEDIA_ROOT, 'uploads/categories/default.png')
                if os.path.exists(default_path):
                    with open(default_path, 'rb') as f:
                        image = ImageFile(f, name='default.png')
                    messages.warning(request, f"Image '{image_name}' not found. Used default image.")
                else:
                    image = None
                    messages.warning(request, f"Image '{image_name}' not found and default image missing.")

            MainCategory.objects.create(
                title=title,
                price=price,
                image=image
            )
            created_count += 1

        messages.success(request, f"{created_count} main categories uploaded successfully.")
        return redirect('custom_admin:main_categories')

    return render(request, 'custom_admin/products/upload_main_categories.html')



@staff_member_required
def upload_subcategories_view(request):
    if request.method == 'POST' and request.FILES.get('excel_file'):
        excel_file = request.FILES['excel_file']
        try:
            df = pd.read_excel(excel_file)
        except Exception as e:
            messages.error(request, f"Error reading Excel file: {e}")
            return redirect('custom_admin:subcategories')

        created_count = 0
        for _, row in df.iterrows():
            main_category_title = str(row.get('Main Category Title')).strip()
            title = str(row.get('Title')).strip()
            price = row.get('Price') or 0.0
            name = str(row.get('Name') or '').strip()
            image_name = str(row.get('Image')).strip() if 'Image' in row else ''

            # Get the MainCategory
            try:
                main_category = MainCategory.objects.get(title__iexact=main_category_title)
            except MainCategory.DoesNotExist:
                messages.warning(request, f"Main category '{main_category_title}' not found. Skipping subcategory '{title}'.")
                continue

            # Try to find uploaded image in MediaLibrary
            media_file = MediaLibrary.objects.filter(name__iexact=image_name).first()
            if media_file:
                image = ImageFile(media_file.file, name=media_file.file.name)
            else:
                # Fallback to default image
                default_path = os.path.join(settings.MEDIA_ROOT, 'uploads/categories/default.png')
                if os.path.exists(default_path):
                    with open(default_path, 'rb') as f:
                        image = ImageFile(f, name='default.png')
                    messages.warning(request, f"Image '{image_name}' not found. Used default image for subcategory '{title}'.")
                else:
                    image = None
                    messages.warning(request, f"Image '{image_name}' not found and default missing for '{title}'.")

            SubCategory.objects.create(
                main_category=main_category,
                title=title,
                name=name,
                price=price,
                image=image
            )
            created_count += 1

        messages.success(request, f"{created_count} subcategories uploaded successfully.")
        return redirect('custom_admin:subcategories')

    return render(request, 'custom_admin/products/upload_subcategories.html')



@staff_member_required
def product_list_view(request):
    products = Product.objects.all()

    # --- Search Filter ---
    query = request.GET.get('q')
    if query:
        products = products.filter(
            Q(name__icontains=query) |
            Q(sku__icontains=query) |
            Q(main_category__title__icontains=query) # Search by related MainCategory title
        )

    # --- Main Category Filter (Dropdown) ---
    main_category_filter = request.GET.get('main_category')
    if main_category_filter:
        products = products.filter(main_category__id=main_category_filter)

    # --- Product Status Filter (Dropdown) ---
    # Assuming 'product_status' is the field on your Product model
    # and it uses choices (e.g., STATUS_CHOICES = [('active', 'Active'), ('inactive', 'Inactive')])
    product_status_filter = request.GET.get('status')
    if product_status_filter:
        products = products.filter(product_status=product_status_filter)

    # --- Sort Filter ---
    sort_by = request.GET.get('sort_by', 'name') # Default sort by product name
    order = request.GET.get('order', 'asc') # Default order ascending

    # Validate sort_by field to prevent arbitrary lookups
    valid_sort_fields = ['name', 'sku', 'main_category__title', 'product_status']
    if sort_by in valid_sort_fields:
        if order == 'desc':
            products = products.order_by(f'-{sort_by}')
        else:
            products = products.order_by(sort_by)
    else:
        # Default sort if an invalid sort_by is provided
        products = products.order_by('name')


    # Get data for dropdown filters
    main_categories_for_filter = MainCategory.objects.all().order_by('title')
    # Assuming your Product model has a product_status field with choices
    product_status_choices = Product._meta.get_field('product_status').choices


    # Determine if any filters are active for the "No results" message
    filters_active = bool(query or main_category_filter or product_status_filter)


    context = {
        'products': products,
        'query': query,
        'main_category_filter': main_category_filter,
        'product_status_filter': product_status_filter,
        'sort_by': sort_by,
        'order': order,
        'main_categories_for_filter': main_categories_for_filter,
        'product_status_choices': product_status_choices,
        'filters_active': filters_active, # Pass the flag to the template
    }
    return render(request, 'custom_admin/products/product_list.html', context)

def add_product_view(request):
    if request.method == 'POST':
        form = ProductForm(request.POST, request.FILES)

        if form.is_valid():
            with transaction.atomic():  # Ensure atomicity
                product = form.save()
            return redirect('custom_admin:product_list')
        else:
            print(form.errors)  # Debug errors
    else:
        form = ProductForm()

    return render(
        request,
        'custom_admin/products/add_product.html',
        {'form': form}
    )


@staff_member_required
def edit_product_view(request, pk):
    product = get_object_or_404(Product, pk=pk)

    # Initialize the ProductForm with the product instance
    if request.method == 'POST':
        form = ProductForm(request.POST, request.FILES, instance=product)

        if form.is_valid():
            form.save()
            return redirect('custom_admin:product_list')  # Redirect to your product list page
    else:
        form = ProductForm(instance=product)

    context = {
        'form': form,
        'product': product,  # Pass product to template if needed
    }
    return render(request, 'custom_admin/products/edit_product.html', context)  # Assuming this is your edit template
def delete_product_view(request, pk):
    product = get_object_or_404(Product, pk=pk)
    if request.method == 'POST':
        product.delete()
        return redirect('custom_admin:product_list')
    return render(request, 'custom_admin/products/delete_product.html', {'product': product})


@staff_member_required
def variant_list_view(request):
    # Start with all variants and prefetch related size options for efficiency
    variants = ProductVariant.objects.all().prefetch_related('size_options')

    # --- Apply Annotations for Total Stock Always ---
    # Annotate total_stock_display for all variants. This makes it available regardless of sorting.
    variants = variants.annotate(
        total_stock_display=Sum('size_options__stock_quantity')
    )

    # --- Search Filter ---
    query = request.GET.get('q')
    if query:
        variants = variants.filter(
            Q(vid__icontains=query) |
            Q(product__name__icontains=query) |
            Q(color__icontains=query)
        )

    # --- Product Filter (Dropdown) ---
    product_filter = request.GET.get('product')
    if product_filter:
        variants = variants.filter(product__id=product_filter)

    # --- Gender Filter (Dropdown) ---
    gender_filter = request.GET.get('gender')
    if gender_filter:
        variants = variants.filter(gender=gender_filter)

    # --- Sort Filter ---
    sort_by = request.GET.get('sort_by', 'vid') # Default sort by Variant ID
    order = request.GET.get('order', 'asc') # Default order ascending

    valid_sort_fields = ['vid', 'product__name', 'color', 'gender', 'total_stock_display'] # Use the annotated field name
    # We now sort by 'total_stock_display' directly since it's always annotated.
    if sort_by in valid_sort_fields:
        if order == 'desc':
            variants = variants.order_by(f'-{sort_by}')
        else:
            variants = variants.order_by(sort_by)
    else:
        variants = variants.order_by('vid') # Default sort if invalid field

    # Get data for dropdown filters
    products_for_filter = Product.objects.all().order_by('name')
    gender_choices = ProductVariant._meta.get_field('gender').choices

    # Determine if any filters are active for the "No results" message
    filters_active = bool(query or product_filter or gender_filter)

    context = {
        'variants': variants,
        'query': query,
        'product_filter': product_filter,
        'gender_filter': gender_filter,
        'sort_by': sort_by,
        'order': order,
        'products_for_filter': products_for_filter,
        'gender_choices': gender_choices,
        'filters_active': filters_active,
    }
    return render(request, 'custom_admin/products/variant_list.html', context)


@staff_member_required
def upload_products_view(request):
    if request.method == 'POST' and request.FILES.get('excel_file'):
        excel_file = request.FILES['excel_file']

        try:
            df = pd.read_excel(excel_file)
        except Exception as e:
            messages.error(request, f"Failed to read Excel file: {e}")
            return redirect('custom_admin:product_list')

        created_count = 0
        error_logs = []

        for index, row in df.iterrows():
            try:
                # Read both possible header variants
                main_category_title = str(row.get('Main Category Title') or row.get('Main Category') or '').strip()
                sub_category_title = str(row.get('Subcategory Title') or row.get('Sub Category') or '').strip()

                # DEBUG: Show row data in console
                print(f"Processing Row {index + 2}: {row.to_dict()}")
                print(f"Extracted Main Category: {main_category_title}, Sub Category: {sub_category_title}")

                if not main_category_title:
                    error_logs.append(f"Row {index + 2}: Missing Main Category.")
                    continue

                if not sub_category_title:
                    error_logs.append(f"Row {index + 2}: Missing Sub Category.")
                    continue

                # Fetch related objects
                try:
                    main_category = MainCategory.objects.get(title__iexact=main_category_title)
                except MainCategory.DoesNotExist:
                    error_logs.append(f"Row {index + 2}: Main category '{main_category_title}' not found.")
                    continue

                try:
                    sub_category = SubCategory.objects.get(title__iexact=sub_category_title, main_category=main_category)
                except SubCategory.DoesNotExist:
                    error_logs.append(f"Row {index + 2}: Subcategory '{sub_category_title}' not found under '{main_category_title}'.")
                    continue

                # Create Product
                product = Product.objects.create(
                    user=request.user,
                    name=row.get('Product Name', '').strip(),
                    title=row.get('Product Title', '').strip(),
                    description=row.get('Description', '').strip(),
                    specification=row.get('Specification', '').strip(),
                    main_category=main_category,
                    sub_category=sub_category,
                    product_status=row.get('Product Status', '').strip(),
                    status=bool(int(row.get('Status (1/0)', 0))),
                    in_stock=bool(int(row.get('In Stock (1/0)', 0))),
                    featured=bool(int(row.get('Featured (1/0)', 0))),
                )

                created_count += 1

            except Exception as e:
                error_logs.append(f"Row {index + 2}: Unexpected error - {e}")

        # Log all messages
        if error_logs:
            for log in error_logs:
                messages.warning(request, log)

        if created_count > 0:
            messages.success(request, f"{created_count} products uploaded successfully.")
        else:
            messages.error(request, "No products were added. Please check your file for errors.")

        return redirect('custom_admin:product_list')

    return render(request, 'custom_admin/products/upload_products.html')

def add_variant_view(request):
    if request.method == 'POST':
        form = ProductVariantForm(request.POST, request.FILES)
        size_formset = VariantSizeOptionFormSet(request.POST)
        image_formset = VariantExtraImageFormSet(request.POST, request.FILES)

        if form.is_valid() and size_formset.is_valid() and image_formset.is_valid():
            variant = form.save()
            size_formset.instance = variant
            size_formset.save()
            image_formset.instance = variant
            image_formset.save()
            return redirect('custom_admin:variant_list')
    else:
        form = ProductVariantForm()
        size_formset = VariantSizeOptionFormSet()
        image_formset = VariantExtraImageFormSet()

    return render(request, 'custom_admin/products/add_variant.html', {
        'form': form,
        'size_formset': size_formset,
        'image_formset': image_formset
    })


@staff_member_required
def edit_variant_view(request, pk):
    variant = get_object_or_404(ProductVariant, pk=pk)

    if request.method == 'POST':
        form = ProductVariantForm(request.POST, request.FILES, instance=variant)
        size_formset = VariantSizeOptionFormSet(request.POST, instance=variant)
        image_formset = VariantExtraImageFormSet(request.POST, request.FILES, instance=variant)

        if form.is_valid() and size_formset.is_valid() and image_formset.is_valid():
            form.save()
            size_formset.save()
            image_formset.save()
            return redirect('custom_admin:variant_list')
    else:
        form = ProductVariantForm(instance=variant)
        size_formset = VariantSizeOptionFormSet(instance=variant)
        image_formset = VariantExtraImageFormSet(instance=variant)

    return render(request, 'custom_admin/products/edit_variant.html', {
        'form': form,
        'size_formset': size_formset,
        'image_formset': image_formset,
        'variant': variant,
    })



def delete_variant_view(request, pk):
    variant = get_object_or_404(ProductVariant, pk=pk)
    if request.method == 'POST':
        variant.delete()
        return redirect('custom_admin:variant_list')
    return render(request, 'custom_admin/products/delete_variant.html', {'variant': variant})

@staff_member_required
def upload_variants_view(request):
    if request.method == 'POST' and request.FILES.get('excel_file'):
        excel_file = request.FILES['excel_file']
        try:
            df = pd.read_excel(excel_file)
        except Exception as e:
            messages.error(request, f"Failed to read Excel file: {e}")
            return redirect('custom_admin:variant_list')

        created_count = 0
        error_logs = []
        grouped = defaultdict(list)

        for index, row in df.iterrows():
            try:
                sku = str(row.get('Product SKU')).strip()
                color = str(row.get('Color')).strip()
                gender = str(row.get('Gender')).strip()
                image_filename = str(row.get('Variant Image') or '').strip()

                # Collect extra image fields
                extra_images = []
                for i in range(1, 7):
                    extra_field = f'Extra Image {i}'
                    img_name = str(row.get(extra_field)).strip() if row.get(extra_field) else ''
                    if img_name:
                        extra_images.append(img_name)

                size = str(row.get('Size')).strip()
                price = row.get('Price')
                old_price = row.get('Old Price', None)
                stock_quantity = int(row.get('Stock Quantity', 0))

                key = (sku, color, gender)
                grouped[key].append({
                    'size': size,
                    'price': price,
                    'old_price': old_price,
                    'stock_quantity': stock_quantity,
                    'variant_image': image_filename,
                    'extra_images': extra_images,
                    'row_index': index + 2
                })
            except Exception as e:
                error_logs.append(f"Row {index + 2}: Error parsing data - {e}")

        for (sku, color, gender), size_rows in grouped.items():
            try:
                product = Product.objects.get(sku__iexact=sku)
            except Product.DoesNotExist:
                error_logs.append(f"Product with SKU '{sku}' not found.")
                continue

            # Load variant image from first row
            variant_image_filename = size_rows[0].get('variant_image')
            image_path = os.path.join(settings.MEDIA_ROOT, 'uploads/media_library', variant_image_filename)
            dummy_path = os.path.join(settings.MEDIA_ROOT, 'variants/dummy.jpg')

            if variant_image_filename and os.path.exists(image_path):
                with open(image_path, 'rb') as img_file:
                    variant_image = ImageFile(io.BytesIO(img_file.read()), name=variant_image_filename)
            elif os.path.exists(dummy_path):
                with open(dummy_path, 'rb') as img_file:
                    variant_image = ImageFile(io.BytesIO(img_file.read()), name='dummy.jpg')
            else:
                error_logs.append(f"Image not found for SKU '{sku}' and dummy missing.")
                continue

            # Create ProductVariant
            variant = ProductVariant.objects.create(
                product=product,
                color=color,
                gender=gender,
                image=variant_image
            )

            # Handle extra images
            extra_image_names = size_rows[0].get('extra_images', [])
            for img_name in extra_image_names:
                path = os.path.join(settings.MEDIA_ROOT, 'uploads/media_library', img_name)
                if os.path.exists(path):
                    with open(path, 'rb') as img_file:
                        extra_image_file = ImageFile(io.BytesIO(img_file.read()), name=img_name)
                        VariantExtraImage.objects.create(product_variant=variant, image=extra_image_file)
                else:
                    error_logs.append(f"Extra image '{img_name}' not found for SKU '{sku}'.")

            # Create all size options
            for size_data in size_rows:
                VariantSizeOption.objects.create(
                    variant=variant,
                    size=size_data['size'],
                    price=size_data['price'],
                    old_price=size_data['old_price'] if not pd.isna(size_data['old_price']) else None,
                    stock_quantity=size_data['stock_quantity']
                )
                created_count += 1

        if error_logs:
            for log in error_logs:
                messages.warning(request, log)

        if created_count:
            messages.success(request, f"{created_count} size options uploaded under multiple variants.")
        else:
            messages.error(request, "No variants were created. Check the Excel file for issues.")

        return redirect('custom_admin:variant_list')

    return render(request, 'custom_admin/products/upload_variants.html')

def generate_variant_label(request, variant_id, size):
    from products.models import ProductVariant, VariantSizeOption

    try:
        variant = ProductVariant.objects.get(vid=variant_id)
        size_option = variant.size_options.get(size=size)
    except (ProductVariant.DoesNotExist, VariantSizeOption.DoesNotExist):
        return HttpResponse("Variant or size not found.", status=404)

    buffer = BytesIO()
    c = canvas.Canvas(buffer, pagesize=(70 * mm, 38 * mm))

    # Header
    c.setFont("Helvetica-Bold", 8)
    c.drawString(10, 95, "BAG LABEL")
    c.drawString(120, 95, "CONVERSION")

    # Basic Info
    c.setFont("Helvetica", 7)
    c.drawString(10, 85, variant.vid.upper())
    c.drawString(120, 85, "BAG")
    c.drawString(10, 70, "2274")
    c.drawString(160, 70, "11118454")

    # Draw Barcode (directly on canvas)
    barcode = code128.Code128(variant.vid.upper(), barHeight=20, barWidth=0.6)
    barcode.drawOn(c, 10, 45)

    # SKU, Price, Color, Size
    c.setFont("Helvetica", 7)
    c.drawString(10, 35, f"M.R.P. ₹ :- {size_option.price}")
    c.drawString(100, 35, "COLOUR")
    c.drawString(160, 35, "SIZE")
    c.drawString(10, 25, "Pkt. Dt. Apr-24")
    c.drawString(100, 25, variant.color.upper())
    c.drawString(160, 25, size.upper())

    c.showPage()
    c.save()
    buffer.seek(0)

    return FileResponse(buffer, as_attachment=True, filename=f"{variant.vid}_{size}.pdf")


def generate_all_variant_labels(request):
    buffer = BytesIO()
    zip_buffer = zipfile.ZipFile(buffer, 'w', zipfile.ZIP_DEFLATED)

    for variant in ProductVariant.objects.prefetch_related('size_options').all():
        for size_option in variant.size_options.all():
            # Generate PDF
            pdf_stream = BytesIO()
            c = canvas.Canvas(pdf_stream, pagesize=(70 * mm, 38 * mm))

            c.setFont("Helvetica-Bold", 8)
            c.drawString(10, 95, "BAG LABEL")
            c.drawString(120, 95, "CONVERSION")

            c.setFont("Helvetica", 7)
            c.drawString(10, 85, variant.vid.upper())
            c.drawString(120, 85, "BAG")
            c.drawString(10, 70, "2274")
            c.drawString(160, 70, "11118454")

            # Barcode
            barcode = code128.Code128(variant.vid.upper(), barHeight=20, barWidth=0.6)
            barcode.drawOn(c, 10, 45)

            # SKU, Price, Color, Size
            c.drawString(10, 35, f"M.R.P. ₹ :- {size_option.price}")
            c.drawString(100, 35, "COLOUR")
            c.drawString(160, 35, "SIZE")
            c.drawString(10, 25, "Pkt. Dt. Apr-24")
            c.drawString(100, 25, variant.color.upper())
            c.drawString(160, 25, size_option.size.upper())

            c.showPage()
            c.save()
            pdf_stream.seek(0)

            # Add to zip
            filename = f"{variant.vid}_{size_option.size}.pdf"
            zip_buffer.writestr(filename, pdf_stream.getvalue())

    zip_buffer.close()
    buffer.seek(0)

    response = HttpResponse(buffer, content_type='application/zip')
    response['Content-Disposition'] = 'attachment; filename=all_variant_tags.zip'
    return response
# Order List View


def order_list_view(request):
    # Start with all orders
    orders = Order.objects.select_related('user').all() # Optimized with select_related

    # Retrieve filter parameters from the request's GET query string
    search_query = request.GET.get('q')
    status_filter = request.GET.get('status')
    start_date_str = request.GET.get('start_date')
    end_date_str = request.GET.get('end_date')

    # Apply the search query filter (Order ID or Username)
    if search_query:
        orders = orders.filter(
            Q(order_id__icontains=search_query) |
            Q(user__username__icontains=search_query)
        )

    # Apply the status filter
    if status_filter:
        orders = orders.filter(status=status_filter)

    # Apply the date range filters
    if start_date_str:
        try:
            # ✅ CORRECTED: Removed the extra '.datetime'
            start_date_obj = datetime.strptime(start_date_str, '%Y-%m-%d').date()
            orders = orders.filter(created_at__date__gte=start_date_obj)
        except (ValueError, TypeError):
            pass

    if end_date_str:
        try:
            # ✅ CORRECTED: Removed the extra '.datetime'
            end_date_obj = datetime.strptime(end_date_str, '%Y-%m-%d').date()
            orders = orders.filter(created_at__date__lte=end_date_obj)
        except (ValueError, TypeError):
            pass

    # Calculate final_total for each order before passing to the template.
    for order in orders:
        order.final_total = order.total - (order.discount if order.discount is not None else 0)

    statuses = Order.STATUS_CHOICES

    context = {
        'orders': orders,
        'statuses': statuses,
        # Pass back the filter values so they can persist in the form fields
        'selected_status': status_filter,
        'selected_start_date': start_date_str,
        'selected_end_date': end_date_str,
        'search_query': search_query,
    }
    return render(request, 'custom_admin/order/order_list.html', context)


def add_order_view(request):
    if request.method == 'POST':
        form = OrderForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('custom_admin:order_list')
    else:
        form = OrderForm()
    return render(request, 'custom_admin/order/add_order.html', {'form': form})


# Edit Order View
def edit_order_view(request, pk):
    order = get_object_or_404(Order, pk=pk)
    if request.method == 'POST':
        form = OrderForm(request.POST, instance=order)
        if form.is_valid():
            form.save()
            return redirect('custom_admin:order_list')
    else:
        form = OrderForm(instance=order)
    return render(request, 'custom_admin/order/edit_order.html', {'form': form, 'order': order})


# Delete Order View
def delete_order_view(request, pk):
    order = get_object_or_404(Order, pk=pk)
    if request.method == 'POST':
        order.delete()
        return redirect('custom_admin:order_list')
    return render(request, 'custom_admin/order/delete_order.html', {'order': order})


def update_order_status_view(request, pk):
    # Ensure this is a POST request and the user is a staff member for security
    if request.method == 'POST' and request.user.is_staff:
        order = get_object_or_404(Order, pk=pk)
        new_status = request.POST.get('status')

        # Get the status before making any changes
        old_status = order.status

        # Check if the new status is a valid choice
        valid_statuses = [choice[0] for choice in Order.STATUS_CHOICES]

        if new_status in valid_statuses:
            # *** IMPORTANT: Only update and send email if the status has changed ***
            if new_status != old_status:
                order.status = new_status
                order.save()

                # Call the function to send the email notification
                send_order_status_update_email(order)

                messages.success(request,
                                 f"Order #{order.order_id} status updated to '{order.get_status_display()}' and a notification has been sent to the user.")
            else:
                messages.info(request,
                              f"Order #{order.order_id} status was already '{order.get_status_display()}'. No changes made.")
        else:
            messages.error(request, "Invalid status selected.")

    # Redirect back to the order list page
    return redirect('custom_admin:order_list')

def send_order_status_update_email(order):
    """
    Sends an email to the user when their order status is updated.
    """
    if not order or not order.user:
        return

    user = order.user
    subject = f"Update on your Stellars Order #{order.order_id}"

    # Prepare context for the email template
    context = {
        'user': user,
        'order': order,
    }

    # Render the HTML email content
    html_message = render_to_string('custom_admin/order_status_update_email.html', context)

    # Send the email
    try:
        send_mail(
            subject,
            '',  # Plain text message (optional, Django can use the HTML)
            settings.DEFAULT_FROM_EMAIL,
            [user.email],
            html_message=html_message,
            fail_silently=False,
        )
    except Exception as e:
        # Log the error if email sending fails
        print(f"Error sending email for order {order.order_id}: {e}")


def cart_list_view(request):
    # Start by getting only carts with at least one item
    carts = Cart.objects.annotate(
        item_count=Count('cartitem')
    ).filter(item_count__gt=0).select_related('user')

    # --- Apply Filters ---
    user_query = request.GET.get('user')
    date_query = request.GET.get('date')

    if user_query:
        carts = carts.filter(user__username__icontains=user_query)

    if date_query:
        try:
            # ✅ CORRECTED: Removed the extra '.datetime' to match the import style
            filter_date = datetime.strptime(date_query, '%Y-%m-%d').date()
            carts = carts.filter(updated__date=filter_date)
        except (ValueError, TypeError):
            pass

    # --- Apply Sorting ---
    order_by = request.GET.get('order_by', 'updated')
    order_direction = request.GET.get('order_direction', 'desc')

    valid_sort_fields = ['id', 'user__username', 'total', 'updated']
    if order_by in valid_sort_fields:
        if order_direction == 'desc':
            order_by = f'-{order_by}'
        carts = carts.order_by(order_by)

    context = {
        'carts': carts,
        'request': request,
    }
    return render(request, 'custom_admin/order/cart_list.html', context)

def add_cart_view(request):
    if request.method == 'POST':
        form = CartForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('custom_admin:cart_list')
    else:
        form = CartForm()
    return render(request, 'custom_admin/order/add_cart.html', {'form': form})


def edit_cart_view(request, pk):
    cart = get_object_or_404(Cart, pk=pk)
    if request.method == 'POST':
        form = CartForm(request.POST, instance=cart)
        if form.is_valid():
            form.save()
            return redirect('custom_admin:cart_list')
    else:
        form = CartForm(instance=cart)
    return render(request, 'custom_admin/order/edit_cart.html', {'form': form, 'cart': cart})


def delete_cart_view(request, pk):
    cart = get_object_or_404(Cart, pk=pk)
    if request.method == 'POST':
        cart.delete()
        return redirect('custom_admin:cart_list')
    return render(request, 'custom_admin/order/delete_cart.html', {'cart': cart})


def wishlist_list_view(request):
    # Use select_related for better performance
    wishlists = Wishlist.objects.select_related('user', 'product', 'product_variant').all()

    # Get filter parameters from the request
    user_query = request.GET.get('user')
    date_query = request.GET.get('date')

    # Apply filters
    if user_query:
        wishlists = wishlists.filter(user__username__icontains=user_query)

    if date_query:
        try:
            # ✅ CORRECTED: Removed the extra '.datetime' to match your import style
            filter_date = datetime.strptime(date_query, '%Y-%m-%d').date()
            wishlists = wishlists.filter(date__date=filter_date)
        except (ValueError, TypeError):
            pass

    # Get sorting parameters
    order_by = request.GET.get('order_by', 'date')
    order_direction = request.GET.get('order_direction', 'desc')

    # Apply sorting
    if order_by in ['id', 'user__username', 'product__name', 'date']:
        if order_direction == 'desc':
            order_by = f'-{order_by}'
        wishlists = wishlists.order_by(order_by)

    context = {
        'wishlists': wishlists,
        'request': request,
    }
    return render(request, 'custom_admin/order/wishlist_list.html', context)

# Add View
def add_wishlist_view(request):
    if request.method == 'POST':
        form = WishlistForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('custom_admin:wishlist_list')
    else:
        form = WishlistForm()
    return render(request, 'custom_admin/order/add_wishlist.html', {'form': form})

# Edit View
def edit_wishlist_view(request, pk):
    wishlist = get_object_or_404(Wishlist, pk=pk)
    if request.method == 'POST':
        form = WishlistForm(request.POST, instance=wishlist)
        if form.is_valid():
            form.save()
            return redirect('custom_admin:wishlist_list')
    else:
        form = WishlistForm(instance=wishlist)
    return render(request, 'custom_admin/order/edit_wishlist.html', {'form': form, 'wishlist': wishlist})


# Delete View
def delete_wishlist_view(request, pk):
    wishlist = get_object_or_404(Wishlist, pk=pk)
    if request.method == 'POST':
        wishlist.delete()
        return redirect('custom_admin:wishlist_list')
    return render(request, 'custom_admin/order/delete_wishlist.html', {'wishlist': wishlist})


def user_addresses_view(request, user_id):
    # Fetch user
    user = get_object_or_404(User, id=user_id)

    # Fetch user's shipping and billing addresses
    shipping_addresses = ShippingAddress.objects.filter(user=user)
    billing_addresses = BillingAddress.objects.filter(user=user)

    context = {
        'user': user,
        'shipping_addresses': shipping_addresses,
        'billing_addresses': billing_addresses
    }

    return render(request, 'custom_admin/order/user_addresses.html', context)


@staff_member_required # Add this decorator for admin views
def blog_list_view(request):
    blogs = Blog.objects.all()

    # --- Search Filter ---
    query = request.GET.get('q')
    if query:
        blogs = blogs.filter(
            Q(title__icontains=query) |
            Q(author__icontains=query) # Assuming 'author' is a CharField or similar for text search
        )

    # --- Sort Filter ---
    sort_by = request.GET.get('sort_by', 'created_at') # Default sort by created_at
    order = request.GET.get('order', 'desc') # Default order descending (newest first)

    valid_sort_fields = ['title', 'author', 'created_at']
    if sort_by in valid_sort_fields:
        if order == 'desc':
            blogs = blogs.order_by(f'-{sort_by}')
        else:
            blogs = blogs.order_by(sort_by)
    else:
        blogs = blogs.order_by('-created_at') # Default sort if invalid field

    # Determine if any filters are active for the "No results" message
    filters_active = bool(query) # Only query for now, no other specific filters

    context = {
        'blogs': blogs,
        'query': query,
        'sort_by': sort_by,
        'order': order,
        'filters_active': filters_active,
    }
    return render(request, 'custom_admin/order/blog_list.html', context)
def add_blog_view(request):
    if request.method == 'POST':
        form = BlogForm(request.POST, request.FILES)
        if form.is_valid():
            print("Form is valid, attempting to save...") # Add this
            try:
                form.save()
                print("Blog saved successfully!") # Add this
                return redirect('custom_admin:blog_list')
            except Exception as e:
                print(f"Error saving blog: {e}") # Add this
                # Log the full traceback for production
                import traceback
                traceback.print_exc()
                # You might want to render the form again with an error message
                # return render(request, 'custom_admin/order/add_blog.html', {'form': form, 'error_message': 'An error occurred during save.'})
        else:
            print("Form is NOT valid. Errors:", form.errors) # Add this
    else:
        form = BlogForm()
    return render(request, 'custom_admin/order/add_blog.html', {'form': form})


def edit_blog_view(request, pk):
    blog = get_object_or_404(Blog, pk=pk)
    if request.method == 'POST':
        form = BlogForm(request.POST, request.FILES, instance=blog)
        if form.is_valid():
            form.save()
            return redirect('custom_admin:blog_list')
    else:
        form = BlogForm(instance=blog)
    return render(request, 'custom_admin/order/edit_blog.html', {'form': form, 'blog': blog})

@require_POST
def delete_blog_view(request, pk):
    blog = get_object_or_404(Blog, pk=pk)
    blog.delete()
    return redirect('custom_admin:blog_list')


@staff_member_required
def banner_list_view(request):
    banners = BannerImage.objects.all()

    # --- Search Filter ---
    query = request.GET.get('q')
    if query:
        banners = banners.filter(Q(title__icontains=query))

    # --- Status Filter (Dropdown) ---
    status_filter = request.GET.get('status')
    if status_filter:
        if status_filter == 'active':
            banners = banners.filter(is_active=True)
        elif status_filter == 'inactive':
            banners = banners.filter(is_active=False)

    # --- Sort Filter ---
    sort_by = request.GET.get('sort_by', 'created_at') # Default sort by created_at
    order = request.GET.get('order', 'desc') # Default order descending (newest first)

    valid_sort_fields = ['title', 'is_active', 'created_at']
    if sort_by in valid_sort_fields:
        if order == 'desc':
            banners = banners.order_by(f'-{sort_by}')
        else:
            banners = banners.order_by(sort_by)
    else:
        banners = banners.order_by('-created_at') # Default sort if invalid field

    # Determine if any filters are active for the "No results" message
    filters_active = bool(query or status_filter)

    context = {
        'banners': banners,
        'query': query,
        'status_filter': status_filter,
        'sort_by': sort_by,
        'order': order,
        'filters_active': filters_active,
        # Pass status choices directly to the template for the dropdown
        'status_choices': [
            ('active', 'Active'),
            ('inactive', 'Inactive')
        ]
    }
    return render(request, 'custom_admin/order/banner_list.html', context)
@staff_member_required
def add_banner_view(request):
    if request.method == 'POST':
        form = BannerImageForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            return redirect('custom_admin:banner_list')
    else:
        form = BannerImageForm()
    return render(request, 'custom_admin/order/add_banner.html', {'form': form})

@staff_member_required
def edit_banner_view(request, pk):
    banner = get_object_or_404(BannerImage, pk=pk)
    if request.method == 'POST':
        form = BannerImageForm(request.POST, request.FILES, instance=banner)
        if form.is_valid():
            form.save()
            return redirect('custom_admin:banner_list')
    else:
        form = BannerImageForm(instance=banner)
    return render(request, 'custom_admin/order/edit_banner.html', {'form': form, 'banner': banner})

@staff_member_required
def delete_banner_view(request, pk):
    banner = get_object_or_404(BannerImage, pk=pk)
    if request.method == 'POST':
        banner.delete()
        return redirect('custom_admin:banner_list')
    return render(request, 'custom_admin/order/delete_banner.html', {'banner': banner})


def customer_list_view(request):
    # Get filter and sort parameters from the request
    query = request.GET.get('q', '')
    sort_by = request.GET.get('sort_by', '-date_joined') # Default sort by newest joined
    order = request.GET.get('order', 'desc')

    # Start with the base queryset
    customers = User.objects.filter(is_staff=False).select_related('profile')

    # Apply search filter
    if query:
        customers = customers.filter(
            Q(username__icontains=query) |
            Q(email__icontains=query)
        )

    # Apply sorting
    valid_sort_fields = ['email', 'username', 'last_login', 'date_joined']
    if sort_by in valid_sort_fields:
        if order == 'desc':
            sort_by = f'-{sort_by}'
        customers = customers.order_by(sort_by)

    # Determine if any filters are active to show the "Clear" button
    filters_active = bool(query)

    context = {
        'customers': customers,
        'query': query,
        'sort_by': sort_by,
        'order': order,
        'filters_active': filters_active,
    }
    return render(request, 'custom_admin/order/customer_list.html', context)


def send_reengagement_email(request, user_id):
    """
    Sends a re-engagement email to a user who has been inactive.
    """
    user = get_object_or_404(User, id=user_id)

    subject = f"We've Missed You at Stellars, {user.username}! ✨"
    context = {
        'user': user,
        'cta_link': request.build_absolute_uri('/'),  # Links to your homepage
    }

    # Render the HTML content from the template
    html_message = render_to_string('custom_admin/reengagement_email.html', context)

    try:
        send_mail(
            subject,
            '',  # Plain text message (optional)
            settings.DEFAULT_FROM_EMAIL,
            [user.email],
            html_message=html_message,
            fail_silently=False,
        )
        messages.success(request, f"Re-engagement email sent successfully to {user.username}.")
    except Exception as e:
        messages.error(request, f"Failed to send email: {e}")

    # Redirect back to the customer list page
    return redirect('custom_admin:customer_list')

@staff_member_required
def add_customer_view(request):
    if request.method == 'POST':
        user_form = CustomerForm(request.POST)
        profile_form = ProfileForm(request.POST)

        if user_form.is_valid() and profile_form.is_valid():
            with transaction.atomic():
                user = user_form.save()
                profile = profile_form.save(commit=False)
                profile.user = user
                profile.save()
            return redirect('custom_admin:customer_list')
    else:
        user_form = CustomerForm()
        profile_form = ProfileForm()

    return render(request, 'custom_admin/order/add_customer.html', {'user_form': user_form, 'profile_form': profile_form})

@staff_member_required
def edit_customer_view(request, pk):
    user = get_object_or_404(User, pk=pk)
    profile = user.profile

    if request.method == 'POST':
        user_form = CustomerForm(request.POST, instance=user)
        profile_form = ProfileForm(request.POST, instance=profile)

        if user_form.is_valid() and profile_form.is_valid():
            with transaction.atomic():
                user_form.save()
                profile_form.save()
            return redirect('custom_admin:customer_list')
    else:
        user_form = CustomerForm(instance=user)
        profile_form = ProfileForm(instance=profile)

    return render(request, 'custom_admin/order/edit_customer.html', {'user_form': user_form, 'profile_form': profile_form, 'customer': user})

@staff_member_required
def delete_customer_view(request, pk):
    customer = get_object_or_404(User, pk=pk)
    if request.method == 'POST':
        customer.delete()
        return redirect('custom_admin:customer_list')
    return render(request, 'custom_admin/order/delete_customer.html', {'customer': customer})



@staff_member_required
def about_list_view(request):
    about_items = About.objects.all()
    return render(request, 'custom_admin/order/about_list.html', {'about_items': about_items})


@staff_member_required
def add_about_view(request):
    if request.method == 'POST':
        form = AboutForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            return redirect('custom_admin:about_list')
    else:
        form = AboutForm()

    return render(request, 'custom_admin/order/add_about.html', {'form': form})


@staff_member_required
def edit_about_view(request, pk):
    about_item = get_object_or_404(About, pk=pk)
    if request.method == 'POST':
        form = AboutForm(request.POST, request.FILES, instance=about_item)
        if form.is_valid():
            form.save()
            return redirect('custom_admin:about_list')
    else:
        form = AboutForm(instance=about_item)

    return render(request, 'custom_admin/order/edit_about.html', {'form': form, 'about_item': about_item})

@staff_member_required
def delete_about_view(request, pk):
    about_item = get_object_or_404(About, pk=pk)
    if request.method == 'POST':
        about_item.delete()
        return redirect('custom_admin:about_list')
    return render(request, 'custom_admin/order/delete_about.html', {'about_item': about_item})


def user_list_view(request):
    if request.user.role != 'admin':
        return render(request, 'custom_admin/access_denied.html', {
            'message': "You do not have permission to view this page."
        })

    # Get filter values from the request URL
    search_query = request.GET.get('q', '')
    role_filter = request.GET.get('role', '')
    include_users_filter = request.GET.get('include_users') # ✅ Get the new filter value

    # Start with the base queryset, excluding superusers
    users = User.objects.exclude(is_superuser=True)

    # ✅ Add new logic to hide the 'user' role by default
    if not include_users_filter:
        users = users.exclude(role='user')

    # Apply search filter for username OR email
    if search_query:
        users = users.filter(
            Q(username__icontains=search_query) |
            Q(email__icontains=search_query)
        )

    # Apply role dropdown filter
    if role_filter:
        users = users.filter(role=role_filter)

    role_choices = User._meta.get_field('role').choices

    context = {
        'users': users.order_by('username'),
        'role_choices': role_choices,
        'current_search': search_query,
        'current_role': role_filter,
        'include_users_checked': include_users_filter, # ✅ Pass the checkbox state to the template
    }
    return render(request, 'custom_admin/user_list.html', context)

@login_required
def edit_user_role(request, user_id):
    if request.user.role != 'admin':
        return render(request, 'custom_admin/access_denied.html', {
            'message': "You do not have permission to access this functionality."
        })

    user = get_object_or_404(User, id=user_id)
    form = UserRoleUpdateForm(request.POST or None, instance=user)

    if request.method == 'POST':
        if form.is_valid():
            form.save()
            messages.success(request, f"Role updated for {user.email}")
            return redirect('custom_admin:user_list')

    return render(request, 'custom_admin/edit_user_role.html', {'form': form, 'user': user})


@login_required
def profile_view(request):
    user = request.user
    profile = user.profile
    user_form = UserUpdateForm(request.POST or None, instance=user)
    profile_form = ProfileForm(request.POST or None, request.FILES or None, instance=profile)

    if request.method == 'POST':
        if user_form.is_valid() and profile_form.is_valid():
            user_form.save()
            profile_form.save()
            messages.success(request, "Profile updated successfully!")

    return render(request, 'custom_admin/profile.html', {
        'user_form': user_form,
        'profile_form': profile_form,
        'profile': profile
    })

def site_settings_list(request):
    settings = SiteSettings.objects.all()
    return render(request, 'custom_admin/site_settings/list.html', {'settings': settings})

def site_settings_create(request):
    form = SiteSettingsForm(request.POST or None)
    if form.is_valid():
        form.save()
        return redirect('custom_admin:site_settings_list')
    return render(request, 'custom_admin/site_settings/form.html', {'form': form, 'title': 'Create Site Settings'})

def site_settings_update(request, pk):
    instance = get_object_or_404(SiteSettings, pk=pk)
    form = SiteSettingsForm(request.POST or None, instance=instance)
    if form.is_valid():
        form.save()
        return redirect('custom_admin:site_settings_list')
    return render(request, 'custom_admin/site_settings/form.html', {'form': form, 'title': 'Update Site Settings'})

def site_settings_delete(request, pk):
    instance = get_object_or_404(SiteSettings, pk=pk)
    if request.method == "POST":
        instance.delete()
        return redirect('custom_admin:site_settings_list')
    return render(request, 'custom_admin/site_settings/confirm_delete.html', {'object': instance})


def discount_code_list_view(request):
    coupons = DiscountCode.objects.all()
    return render(request, 'custom_admin/discount_code_list.html', {'coupons': coupons})


def coupon_list_view(request):
    coupons = DiscountCode.objects.all()
    return render(request, 'custom_admin/discount_code_list.html', {'coupons': coupons})

def add_coupon_view(request):
    form = DiscountCodeForm(request.POST or None)
    if form.is_valid():
        form.save()
        return redirect('custom_admin:coupon_list')
    return render(request, 'custom_admin/discount_code_form.html', {'form': form, 'title': 'Add Coupon'})

def edit_coupon_view(request, pk):
    coupon = get_object_or_404(DiscountCode, pk=pk)
    form = DiscountCodeForm(request.POST or None, instance=coupon)
    if form.is_valid():
        form.save()
        return redirect('custom_admin:coupon_list')
    return render(request, 'custom_admin/discount_code_form.html', {'form': form, 'title': 'Edit Coupon'})

def delete_coupon_view(request, pk):
    coupon = get_object_or_404(DiscountCode, pk=pk)
    if request.method == 'POST':
        coupon.delete()
        return redirect('custom_admin:coupon_list')
    return render(request, 'custom_admin/confirm_delete.html', {'object': coupon, 'title': 'Delete Coupon'})


@staff_member_required
def media_library_view(request):
    query = request.GET.get('q', '')
    media_files = MediaLibrary.objects.all()

    if query:
        media_files = media_files.filter(name__icontains=query)

    if request.method == 'POST':
        uploaded_files = request.FILES.getlist('files')  # <-- Handle multiple files
        if not uploaded_files:
            messages.error(request, "No files selected.")
        else:
            success_count = 0
            for file in uploaded_files:
                try:
                    media = MediaLibrary(file=file, name=file.name)
                    media.save()
                    success_count += 1
                except Exception as e:
                    messages.warning(request, f"Error uploading '{file.name}': {str(e)}")

            if success_count:
                messages.success(request, f"{success_count} file(s) uploaded successfully.")
            else:
                messages.error(request, "No files were uploaded.")

        return redirect('custom_admin:media_library')

    return render(request, 'custom_admin/media_library.html', {'media_files': media_files})



@staff_member_required
def delete_media_file(request, pk):
    media = get_object_or_404(MediaLibrary, pk=pk)

    file_path = media.file.path
    if os.path.isfile(file_path):
        os.remove(file_path)

    media.delete()
    messages.success(request, "Media file deleted successfully.")
    return redirect('custom_admin:media_library')


@login_required
def site_content_list(request):
    """
    Displays the current SiteContent instance.
    Provides a link to create it if it doesn't exist, or edit if it does.
    """
    site_content = SiteContent.objects.first() # There should only be one instance

    context = {
        'site_content': site_content,
        'page_title': 'Site Content Management',
    }
    return render(request, 'custom_admin/site_content_list.html', context)

@login_required
def site_content_create_or_update(request):
    """
    Handles creation if no SiteContent exists, or updates the existing one.
    This consolidates add/update into a single view for the single instance.
    """
    site_content_instance = SiteContent.objects.first()
    is_new_instance = not site_content_instance

    if request.method == 'POST':
        form = SiteContentForm(request.POST, instance=site_content_instance)
        if form.is_valid():
            form.save()
            messages.success(request, 'Site Content updated successfully!' if not is_new_instance else 'Site Content created successfully!')
            return redirect('custom_admin:site_content_list') # Redirect to the list view
        else:
            messages.error(request, 'Please correct the error below.')
    else:
        form = SiteContentForm(instance=site_content_instance)

    context = {
        'form': form,
        'is_new_instance': is_new_instance,
        'page_title': 'Edit Site Content' if not is_new_instance else 'Create Site Content',
    }
    return render(request, 'custom_admin/site_content_form.html', context)


@login_required
def site_content_delete(request):
    """
    Handles the deletion of the SiteContent instance.
    Requires strong confirmation as there should ideally be only one.
    """
    site_content_instance = SiteContent.objects.first()

    if not site_content_instance:
        messages.warning(request, "No Site Content to delete.")
        return redirect('custom_admin:site_content_list')

    if request.method == 'POST':
        # Add a hidden input or check for a specific POST parameter for confirmation
        if 'confirm_delete' in request.POST:
            site_content_instance.delete()
            messages.success(request, 'Site Content deleted successfully.')
            return redirect('custom_admin:site_content_list')
        else:
            messages.error(request, 'Deletion not confirmed.')
            return redirect('custom_admin:site_content_list') # Or render the confirmation again

    context = {
        'site_content': site_content_instance,
        'page_title': 'Confirm Delete Site Content',
    }
    return render(request, 'custom_admin/site_content_confirm_delete.html', context)


def send_reminder_email(request, user_id, reminder_type):
    """
    Sends a reminder email to a user for their cart or wishlist.
    """
    user = get_object_or_404(User, id=user_id)

    # Determine the subject, template, and items based on the reminder type
    if reminder_type == 'cart':
        cart = Cart.objects.filter(user=user).first()
        if not cart or cart.cartitem_set.count() == 0:
            messages.warning(request, f"{user.username}'s cart is empty. No email sent.")
            return redirect('custom_admin:cart_list')

        subject = "You left some items in your shopping cart! 🛒"
        template_name = 'custom_admin/cart_reminder_email.html'
        items = cart.cartitem_set.all()
        context = {
            'user': user,
            'items': items,
            'cta_link': request.build_absolute_uri('/cart/'),  # Adjust if your cart URL is different
        }

    elif reminder_type == 'wishlist':
        wishlist_items = Wishlist.objects.filter(user=user)
        if not wishlist_items.exists():
            messages.warning(request, f"{user.username}'s wishlist is empty. No email sent.")
            return redirect('custom_admin:wishlist_list')

        subject = "Still thinking about these items? 👀"
        template_name = 'custom_admin/wishlist_reminder_email.html'
        context = {
            'user': user,
            'items': wishlist_items,
            'cta_link': request.build_absolute_uri('/wishlist/'),  # Adjust if your wishlist URL is different
        }
    else:
        messages.error(request, "Invalid reminder type.")
        return redirect('custom_admin:dashboard')

    # Render the HTML content from the template
    html_message = render_to_string(template_name, context)

    try:
        send_mail(
            subject,
            '',  # Plain text message (optional)
            settings.DEFAULT_FROM_EMAIL,
            [user.email],
            html_message=html_message,
            fail_silently=False,
        )
        messages.success(request, f"Reminder email sent successfully to {user.username}.")
    except Exception as e:
        messages.error(request, f"Failed to send email: {e}")

    # Redirect back to the page the admin came from
    return redirect(request.META.get('HTTP_REFERER', 'custom_admin:dashboard'))


@staff_member_required
def brand_list_view(request):
    brands = Brand.objects.all()
    return render(request, 'custom_admin/brand_list.html', {'brands': brands})

@staff_member_required
def add_brand_view(request):
    if request.method == 'POST':
        form = BrandForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            messages.success(request, 'Brand added successfully!')
            return redirect('custom_admin:brand_list')
    else:
        form = BrandForm()
    # Pass the form and title to the new template
    return render(request, 'custom_admin/brand_form.html', {'form': form, 'title': 'Add New Brand'})

@staff_member_required
def edit_brand_view(request, pk):
    brand = get_object_or_404(Brand, pk=pk)
    if request.method == 'POST':
        form = BrandForm(request.POST, request.FILES, instance=brand)
        if form.is_valid():
            form.save()
            messages.success(request, 'Brand updated successfully!')
            return redirect('custom_admin:brand_list')
    else:
        form = BrandForm(instance=brand)
    # Pass the form and a dynamic title to the new template
    return render(request, 'custom_admin/brand_form.html', {'form': form, 'title': f'Edit Brand: {brand.name}'})

@staff_member_required
def delete_brand_view(request, pk):
    brand = get_object_or_404(Brand, pk=pk)
    if request.method == 'POST':
        brand.delete()
        messages.success(request, 'Brand deleted successfully!')
        return redirect('custom_admin:brand_list')
    return render(request, 'custom_admin/confirm_delete.html', {'object': brand})


def admin_login_view(request):
    """
    Handles the login for staff and admin users.
    """
    if request.user.is_authenticated and request.user.is_staff:
        return redirect('custom_admin:dashboard')

    if request.method == 'POST':
        # Use your new custom form here
        form = CustomAdminAuthenticationForm(request, data=request.POST)
        if form.is_valid():
            username = form.cleaned_data.get('username')
            password = form.cleaned_data.get('password')
            user = authenticate(username=username, password=password)
            if user is not None and user.is_staff:
                login(request, user)
                return redirect('custom_admin:dashboard')
            else:
                messages.error(request, "Invalid credentials or you do not have staff permissions.")
    else:
        # And also use it here for GET requests
        form = CustomAdminAuthenticationForm()

    return render(request, 'custom_admin/login.html', {'form': form})